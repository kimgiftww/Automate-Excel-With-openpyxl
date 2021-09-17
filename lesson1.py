import datetime as dt
import openpyxl as xl

# Construct an Excel workbook object
wb = xl.Workbook()

# Construct an Excel worksheet object
ws = wb.active

# insert a single value to the active worksheet
ws['A1'] = 'Cell A1'

# insert a row of values
record = ('1234', 'Student', 'Science')
ws.append(record)

# openpyxl recognizes datetime object
currentTime = dt.datetime.now()
ws['A3'] = currentTime

# If the existing workbook is open, you must close the 
# Excel file for the wb object to save over the existing file
# otherwise, you will run into PermissionError
wb.save('test1.xlsx')
