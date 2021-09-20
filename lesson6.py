import datetime as dt
import openpyxl as xl
from openpyxl.workbook.workbook import Workbook

"""
Adding/Deleting Rows/Columns
"""

wb = xl.Workbook()

ws = wb.create_sheet('Insert Rows Columns')

ws.cell(2, 1, 'A2')
ws.cell(5, 1, 'A5')
ws.cell(5, 3, 'C5')
ws.cell(5, 3, 'C5')

# insert one row (default) before row 2
# this will push row 2 to become row 3
ws.insert_rows(2)

# to insert multiple rows
# set the row count
ws.insert_rows(5, 10)

# same approach can be applied to insert columns
ws.insert_cols(1, 2)

# 2nd run
# delete row/column
ws.delete_rows(1, 2)
ws.delete_cols(1, 2)

wb.save('lesson6.xlsx')