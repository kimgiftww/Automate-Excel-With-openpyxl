import datetime as dt
import openpyxl as xl

"""
Format Datetime And Insert Fomrula
"""

wb = xl.Workbook()
wb = xl.load_workbook('lesson5.xlsx', data_only=True)
wsDatetime = wb.create_sheet('Datatime')

wsDatetime.cell(1, 1, dt.datetime.now())

# format to mm/dd/yyyy
wsDatetime.cell(2, 1, dt.datetime.now())
wsDatetime.cell(2, 1).number_format = 'mm/dd/yyyy'

# insert formula
wsFormula = wb.create_sheet('Formula')
A1 = wsFormula.cell(1, 1, 20)
A2 = wsFormula.cell(2, 1, 30)
wsFormula.cell(3, 1, 
    '=SUM({0}{1}, {2}{3})'.format(A1.column_letter, A1.row, A2.column_letter, A2.row))

# just noted that since openpyxl never evaluates
# a formula, when you read a formula cell,
# it will return as None
wsFormula.cell(3, 1).value

wb.save('lesson5.xlsx')

