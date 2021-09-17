import datetime as dt
import openpyxl as xl
from PIL import ImageColor

wb = xl.Workbook()

# create a worksheet
ws = {}
# insert worksheet at the end
ws['Mar'] = wb.create_sheet('March')

# we can specify the position using zero based array.
# insert worksheet at the penultimate position
# second from the last
ws['Feb'] = wb.create_sheet('February', -1)

# insert worksheet in the first position
ws['Jan'] = wb.create_sheet('Jan', 0)

# rename a worksheet
ws['Jan'].title = 'January'

# to change a worksheet tab color
# Colors must be aRGB hex values (which is a bit inconvient)
ws['Jan'].sheet_properties.tabColor = '1072BA'
ws['Feb'].sheet_properties.tabColor = '84FF92FF'

# to iterate each worksheet
for sheet in wb:
    print(sheet.title)

wb.save('lesson2.xlsx')

ImageColor.getcolor("#23a9dd", "RGBa")