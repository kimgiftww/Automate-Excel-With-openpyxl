import datetime as dt
import openpyxl as xl

wb = xl.Workbook()
ws = wb.active


#########################
#       Write Data       #
#########################

cell_address = 'C3'
ws[cell_address] = 'Cell C3'

# openpyxl uses 1 based array index
row_index = 1
column_index = 1
value = 10
ws.cell(row_index, column_index, value=value)

values = [10, 20, 30]
# ws.append() will always add rows below any existing data.
# unfortunately, openpyxl doesn't give you the option to
# insert an array of values at a specific location.
ws.append(values)

# to add an array of values, we need to iterate each value
# individually.
values = [11, 21, 31]
row_insertion = ws.max_row
column_insertion = 1

for indx, value in enumerate(values):
    ws.cell(row_insertion + 1, column_insertion + indx, value=value)

records = [
    [12, 22, 32],
    [13, 23, 33]
]
row_insertion = ws.max_row

for indx_row, record in enumerate(records):
    for indx_col, value in enumerate(record):
        # print(row_insertion + 1 + indx_row, column_insertion + indx_col)
        ws.cell(row_insertion + 1 + indx_row, column_insertion + indx_col, value=value)

##########################
#       Read Data        #
##########################
last_row = ws.max_row
rng = ws['A1':'D4']
rng2 = ws['A1:D4']
cols = ws['B:C']
# not recommended as it will iterate every single cell
rows = ws['2:4']

# to print the cell values
for row in rng:
    for cell in row:
        print(cell.value)

# we can also use the Worksheet.iter_rows() or Worksheet.iter_cols method
# this method
for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=3):
    for cell in row:
        print('Cell Row: {0}, Col: {1}, Value: {2}'.format(cell.row, cell.column, cell.value))
        print('Cell {0}{1} has value: {2}'.format(cell.column_letter, cell.row, cell.value))



wb.save('lesson3.xlsx')